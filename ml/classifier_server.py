"""
Servidor Flask para Classificação de Incidentes de Segurança
Porta: 5001 (interna, não exposta publicamente)

Metodologia ML:
  - Dataset de TREINAMENTO: incidentes_cybersecurity_2000.xlsx (2000 amostras)
    Usado para treinar o modelo TF-IDF + Naive Bayes.
  - Dataset de AVALIAÇÃO: incidentes_cybersecurity_100.xlsx (100 amostras)
    Usado para avaliar o modelo em produção (conjunto independente, nunca visto no treino).

Endpoints:
  GET  /health       - Status do servidor e modelo
  POST /classify     - Classificar incidente (title + description)
  GET  /metrics      - Métricas do modelo (treino + avaliação)
  GET  /dataset      - Download do dataset de treinamento (base64 + preview)
  GET  /eval-dataset - Download do dataset de avaliação (base64 + preview)
  POST /evaluate     - Avaliar modelo com dataset de avaliação (100 amostras)
  POST /retrain      - Retreinar modelo com novas amostras/categorias
  POST /reload-model - Recarregar modelo do disco sem reiniciar
"""

import os
import json
import base64
import joblib
import numpy as np
import pandas as pd
import threading
import time
import unicodedata
import re
from datetime import datetime, timezone
from flask import Flask, request, jsonify, Response, stream_with_context

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH         = os.path.join(SCRIPT_DIR, "model.pkl")
METRICS_PATH       = os.path.join(SCRIPT_DIR, "metrics.json")
TRAIN_DATASET_PATH = os.path.join(SCRIPT_DIR, "incidentes_cybersecurity_2000.xlsx")
EVAL_DATASET_PATH  = os.path.join(SCRIPT_DIR, "incidentes_cybersecurity_100.xlsx")

app = Flask(__name__)

# ─── Cache de Modelo em Memória ──────────────────────────────────────────────────────────
_model_cache = {"pipeline": None, "loaded_at": None}
_startup_complete = False

def get_cached_pipeline():
    """Retorna o pipeline em cache, carregando se necessário."""
    global _model_cache
    if _model_cache["pipeline"] is None:
        print(f"[Classifier] Carregando modelo de {MODEL_PATH}...")
        try:
            _model_cache["pipeline"] = joblib.load(MODEL_PATH)
            _model_cache["loaded_at"] = datetime.now(timezone.utc).isoformat()
            print("[Classifier] Modelo carregado com sucesso!")
        except Exception as e:
            print(f"[Classifier] ERRO ao carregar modelo: {e}")
            _model_cache["pipeline"] = None
    return _model_cache["pipeline"]

# ─── Carregar Modelo (Lazy Loading) ──────────────────────────────────────────────────────────
print(f"[Classifier] Iniciando com lazy loading de modelo...")
pipeline = None  # Será carregado sob demanda

# ─── Carregar Métricas ────────────────────────────────────────────────────────
try:
    with open(METRICS_PATH, "r", encoding="utf-8") as f:
        metrics = json.load(f)
except Exception:
    metrics = {}

# ─── Mapeamento de Risco ──────────────────────────────────────────────────────
RISK_MAP = {
    "phishing": "high",
    "malware": "critical",
    "brute_force": "high",
    "brute force": "high",
    "ddos": "medium",
    "vazamento_de_dados": "critical",
    "vazamento de dados": "critical",
    "unknown": "low",
}

# ─── Utilitário: Ler Dataset XLSX ─────────────────────────────────────────────
def _load_dataset(path: str) -> tuple[list[str], list[str]]:
    """Lê um arquivo XLSX e retorna (texts, labels). Suporta colunas em PT e EN."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    title_col = next((i for i, h in enumerate(headers) if h in ("title", "titulo", "título")), 0)
    desc_col  = next((i for i, h in enumerate(headers) if h in ("description", "descricao", "descrição")), 1)
    cat_col   = next((i for i, h in enumerate(headers) if h in ("category", "categoria")), 2)
    texts, labels = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        t = str(row[title_col]) if row[title_col] else ""
        d = str(row[desc_col])  if row[desc_col]  else ""
        c = str(row[cat_col])   if row[cat_col]   else "unknown"
        c = c.lower().strip().replace(" ", "_")
        texts.append(f"{t} {d}".lower().strip())
        labels.append(c)
    return texts, labels

# ─── Endpoints ────────────────────────────────────────────────────────────────

@app.route("/reload-model", methods=["POST"])
def reload_model():
    """Recarrega o modelo e as métricas do disco sem reiniciar o servidor."""
    global pipeline, metrics
    try:
        pipeline = joblib.load(MODEL_PATH)
        with open(METRICS_PATH, "r", encoding="utf-8") as f:
            metrics = json.load(f)
        print(f"[Classifier] Modelo recarregado: {metrics.get('training', {}).get('dataset_size')} amostras de treino")
        return jsonify({
            "status": "ok",
            "message": "Modelo recarregado com sucesso",
            "training_dataset_size": metrics.get("training", {}).get("dataset_size"),
            "categories": metrics.get("training", {}).get("categories"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/health", methods=["GET"])
def health():
    global _startup_complete
    # Lazy load do modelo na primeira requisição
    if not _startup_complete:
        get_cached_pipeline()
        _startup_complete = True
    return jsonify({
        "status": "ok",
        "model_loaded": _model_cache["pipeline"] is not None,
        "model_loaded_at": _model_cache["loaded_at"],
        "train_dataset": "incidentes_cybersecurity_2000.xlsx",
        "eval_dataset": "incidentes_cybersecurity_100.xlsx",
        "metrics": metrics,
    })


def classify_fallback():
    """Classifica usando palavras-chave quando o modelo nao esta disponivel."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid JSON body"}), 400
    
    title = str(data.get("title", "")).lower()
    description = str(data.get("description", "")).lower()
    text = f"{title} {description}"
    
    keyword_map = {
        "phishing": ["phishing", "email falso", "link suspeito", "credenciais"],
        "malware": ["malware", "virus", "trojan", "ransomware", "worm"],
        "brute_force": ["brute force", "forca bruta", "senha", "login", "tentativas"],
        "ddos": ["ddos", "negacao de servico", "ataque distribuido", "trafego"],
        "vazamento_de_dados": ["vazamento", "dados", "leak", "exposicao", "roubo"],
    }
    
    scores = {cat: 0 for cat in keyword_map}
    for category, keywords in keyword_map.items():
        for keyword in keywords:
            scores[category] += text.count(keyword)
    
    category = max(scores, key=scores.get) if max(scores.values()) > 0 else "unknown"
    confidence = min(0.6 + (max(scores.values()) * 0.05), 0.95) if max(scores.values()) > 0 else 0.3
    risk_level = RISK_MAP.get(category, "low")
    
    return jsonify({
        "category": category,
        "confidence": round(confidence, 4),
        "risk_level": risk_level,
        "probabilities": {cat: round(scores[cat] / max(sum(scores.values()), 1), 4) for cat in scores},
        "model_info": {
            "train_dataset": "incidentes_cybersecurity_2000.xlsx",
            "fallback_mode": True,
        },
    })

@app.route("/classify", methods=["POST"])
def classify():
    """Classifica um incidente usando o modelo treinado com o dataset de 2000 amostras."""
    pipeline = get_cached_pipeline()
    if pipeline is None:
        # Fallback: classificação por palavras-chave
        return classify_fallback()

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid JSON body"}), 400

    title       = str(data.get("title", "")).strip()
    description = str(data.get("description", "")).strip()

    if not title and not description:
        return jsonify({"error": "title or description required"}), 400

    text = f"{title} {description}".lower().strip()

    try:
        category   = pipeline.predict([text])[0]
        proba      = pipeline.predict_proba([text])[0]
        confidence = float(max(proba))
        risk_level = RISK_MAP.get(category, "low")
        classes    = pipeline.classes_.tolist()
        all_probas = {cls: float(prob) for cls, prob in zip(classes, proba)}

        return jsonify({
            "category":     category,
            "confidence":   round(confidence, 4),
            "risk_level":   risk_level,
            "probabilities": all_probas,
            # Informação metodológica: qual dataset foi usado para treinar o modelo
            "model_info": {
                "train_dataset":      "incidentes_cybersecurity_2000.xlsx",
                "train_dataset_size": metrics.get("training", {}).get("dataset_size", 2000),
            },
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/metrics", methods=["GET"])
def get_metrics():
    """
    Retorna métricas separadas para treino e avaliação.
    Estrutura:
      {
        training: { dataset, dataset_size, train_accuracy, cv_accuracy_mean, ... },
        evaluation: { dataset, dataset_size, eval_accuracy, ... } | null,
        method: "TF-IDF + Naive Bayes",
        categories: [...],
        last_updated: "...",
      }
    """
    return jsonify(metrics)


@app.route("/dataset", methods=["GET"])
def get_dataset():
    """Retorna o dataset de TREINAMENTO (2000 amostras) em base64 + preview."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(TRAIN_DATASET_PATH)
        ws = wb.active
        headers  = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        title_col = next((i for i, h in enumerate(headers) if h in ("title", "titulo", "título")), 0)
        desc_col  = next((i for i, h in enumerate(headers) if h in ("description", "descricao", "descrição")), 1)
        cat_col   = next((i for i, h in enumerate(headers) if h in ("category", "categoria")), 2)

        preview, category_dist, total = [], {}, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            cat = str(row[cat_col]) if row[cat_col] else "unknown"
            category_dist[cat] = category_dist.get(cat, 0) + 1
            total += 1
            if len(preview) < 10:
                preview.append({
                    "title":       str(row[title_col]) if row[title_col] else "",
                    "description": str(row[desc_col])  if row[desc_col]  else "",
                    "category":    cat,
                })

        with open(TRAIN_DATASET_PATH, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")

        return jsonify({
            "role":                 "training",
            "filename":             "incidentes_cybersecurity_2000.xlsx",
            "base64":               b64,
            "total_samples":        total,
            "category_distribution": category_dist,
            "preview":              preview,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/eval-dataset", methods=["GET"])
def get_eval_dataset():
    """Retorna o dataset de AVALIAÇÃO (100 amostras) em base64 + preview."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EVAL_DATASET_PATH)
        ws = wb.active
        headers  = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        title_col = next((i for i, h in enumerate(headers) if h in ("title", "titulo", "título")), 0)
        desc_col  = next((i for i, h in enumerate(headers) if h in ("description", "descricao", "descrição")), 1)
        cat_col   = next((i for i, h in enumerate(headers) if h in ("category", "categoria")), 2)

        preview, category_dist, total = [], {}, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            cat = str(row[cat_col]) if row[cat_col] else "unknown"
            category_dist[cat] = category_dist.get(cat, 0) + 1
            total += 1
            if len(preview) < 10:
                preview.append({
                    "title":       str(row[title_col]) if row[title_col] else "",
                    "description": str(row[desc_col])  if row[desc_col]  else "",
                    "category":    cat,
                })

        with open(EVAL_DATASET_PATH, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")

        return jsonify({
            "role":                 "evaluation",
            "filename":             "incidentes_cybersecurity_100.xlsx",
            "base64":               b64,
            "total_samples":        total,
            "category_distribution": category_dist,
            "preview":              preview,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/evaluate", methods=["POST"])
def evaluate():
    """
    Avalia o modelo usando o dataset de AVALIAÇÃO (100 amostras).
    Este conjunto é independente do treino — nunca foi usado para ajustar o modelo.
    Retorna: accuracy, per-category precision/recall/f1, confusion matrix.
    """
    if pipeline is None:
        return jsonify({"error": "Model not loaded"}), 503

    try:
        from sklearn.metrics import (
            accuracy_score, classification_report,
            confusion_matrix,
        )

        eval_texts, eval_labels = _load_dataset(EVAL_DATASET_PATH)
        if not eval_texts:
            return jsonify({"error": "Dataset de avaliação vazio ou não encontrado"}), 500

        predictions = pipeline.predict(eval_texts)
        accuracy    = float(accuracy_score(eval_labels, predictions))

        # Relatório por categoria
        report_dict = classification_report(
            eval_labels, predictions,
            output_dict=True,
            zero_division=0,
        )

        # Matriz de confusão
        classes = sorted(set(eval_labels))
        cm      = confusion_matrix(eval_labels, predictions, labels=classes).tolist()

        # Distribuição do dataset de avaliação
        eval_dist: dict[str, int] = {}
        for lbl in eval_labels:
            eval_dist[lbl] = eval_dist.get(lbl, 0) + 1

        eval_result = {
            "dataset":              "incidentes_cybersecurity_100.xlsx",
            "dataset_size":         len(eval_texts),
            "eval_accuracy":        round(accuracy, 4),
            "category_distribution": eval_dist,
            "per_category":         {
                k: {
                    "precision": round(v["precision"], 4),
                    "recall":    round(v["recall"], 4),
                    "f1_score":  round(v["f1-score"], 4),
                    "support":   int(v["support"]),
                }
                for k, v in report_dict.items()
                if k not in ("accuracy", "macro avg", "weighted avg")
            },
            "macro_avg": {
                "precision": round(report_dict["macro avg"]["precision"], 4),
                "recall":    round(report_dict["macro avg"]["recall"], 4),
                "f1_score":  round(report_dict["macro avg"]["f1-score"], 4),
            },
            "weighted_avg": {
                "precision": round(report_dict["weighted avg"]["precision"], 4),
                "recall":    round(report_dict["weighted avg"]["recall"], 4),
                "f1_score":  round(report_dict["weighted avg"]["f1-score"], 4),
            },
            "confusion_matrix": {
                "labels": classes,
                "matrix": cm,
            },
            "evaluated_at": datetime.now(timezone.utc).isoformat(),
        }

        # Persistir resultado de avaliação nas métricas
        global metrics
        metrics["evaluation"] = eval_result
        metrics["last_evaluated"] = eval_result["evaluated_at"]
        with open(METRICS_PATH, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=2)

        return jsonify({
            "success":    True,
            "evaluation": eval_result,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/retrain", methods=["POST"])
def retrain():
    """
    Retreina o modelo usando o dataset de TREINAMENTO (2000 amostras)
    mais quaisquer amostras adicionais fornecidas no body.
    O dataset de AVALIAÇÃO (100 amostras) NUNCA é incluído no treino.
    """
    global pipeline, metrics, RISK_MAP

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid JSON body"}), 400

    new_samples  = data.get("samples", [])
    new_risk_map = data.get("risk_map", {})

    # Se samples estiver vazio, retreinar apenas com o dataset atual (ex: após upload de novo dataset)
    # Se samples tiver itens, adicionar ao dataset existente

    for s in new_samples:
        if not s.get("description", "").strip():
            return jsonify({"error": "Each sample must have a non-empty description"}), 400
        if not s.get("category", "").strip():
            return jsonify({"error": "Each sample must have a non-empty category"}), 400

    try:
        from sklearn.pipeline import Pipeline
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.naive_bayes import MultinomialNB
        from sklearn.model_selection import cross_val_score

        # 1. Carregar APENAS o dataset de treinamento (2000 amostras)
        texts, labels = _load_dataset(TRAIN_DATASET_PATH)

        # 2. Adicionar novas amostras (do banco de dados ou manuais)
        new_categories: set[str] = set()
        for s in new_samples:
            t = str(s.get("title", ""))
            d = str(s.get("description", ""))
            c = str(s.get("category", "")).strip().lower().replace(" ", "_")
            texts.append(f"{t} {d}".lower().strip())
            labels.append(c)
            if c not in labels[:-1]:
                new_categories.add(c)

        # 3. Atualizar RISK_MAP
        for cat, risk in new_risk_map.items():
            RISK_MAP[cat.lower().replace(" ", "_")] = risk
        for cat in new_categories:
            if cat not in RISK_MAP:
                RISK_MAP[cat] = "medium"

        # 4. Retreinar o pipeline
        new_pipeline = Pipeline([
            ("tfidf", TfidfVectorizer(ngram_range=(1, 2), max_features=5000, sublinear_tf=True)),
            ("clf",   MultinomialNB()),
        ])
        new_pipeline.fit(texts, labels)

        # 5. Cross-validation (5-fold) no conjunto de treino
        cv_scores = cross_val_score(
            new_pipeline, texts, labels,
            cv=min(5, len(set(labels))), scoring="accuracy",
        )
        train_acc = float(new_pipeline.score(texts, labels))

        # 6. Salvar modelo
        joblib.dump(new_pipeline, MODEL_PATH)
        pipeline = new_pipeline

        # 7. Distribuição de categorias no treino
        cat_dist: dict[str, int] = {}
        for label in labels:
            cat_dist[label] = cat_dist.get(label, 0) + 1

        # 8. Atualizar métricas — preservar avaliação anterior se existir
        training_metrics = {
            "dataset":              "incidentes_cybersecurity_2000.xlsx",
            "dataset_size":         len(texts),
            "categories":           list(new_pipeline.classes_),
            "cv_accuracy_mean":     round(float(np.mean(cv_scores)), 4),
            "cv_accuracy_std":      round(float(np.std(cv_scores)), 4),
            "train_accuracy":       round(train_acc, 4),
            "category_distribution": cat_dist,
            "new_samples_added":    len(new_samples),
            "new_categories":       list(new_categories),
        }

        new_metrics = {
            "method":       "TF-IDF + Naive Bayes (MultinomialNB)",
            "categories":   list(new_pipeline.classes_),
            "training":     training_metrics,
            "evaluation":   metrics.get("evaluation"),  # preservar avaliação anterior
            "last_updated": datetime.now(timezone.utc).isoformat(),
            "last_evaluated": metrics.get("last_evaluated"),
            # Campos legados para compatibilidade com código existente
            "dataset_size":         len(texts),
            "cv_accuracy_mean":     round(float(np.mean(cv_scores)), 4),
            "cv_accuracy_std":      round(float(np.std(cv_scores)), 4),
            "train_accuracy":       round(train_acc, 4),
            "category_distribution": cat_dist,
        }

        with open(METRICS_PATH, "w", encoding="utf-8") as f:
            json.dump(new_metrics, f, ensure_ascii=False, indent=2)
        metrics = new_metrics

        return jsonify({
            "success": True,
            "message": (
                f"Modelo retreinado com {len(texts)} amostras de treino "
                f"({len(new_samples)} novas). "
                f"Dataset de avaliação (100 amostras) preservado separadamente."
            ),
            "metrics": new_metrics,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Upload de Dataset de Treinamento ──────────────────────────────────────────
@app.route("/upload-train-dataset", methods=["POST"])
def upload_train_dataset():
    """Substitui o dataset de TREINAMENTO por um novo arquivo .xlsx enviado pelo admin."""
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Apenas arquivos .xlsx são aceitos"}), 400
    try:
        # Salvar o novo arquivo substituindo o dataset de treino
        f.save(TRAIN_DATASET_PATH)
        # Verificar quantas amostras o novo dataset tem usando pandas
        df_new = pd.read_excel(TRAIN_DATASET_PATH)
        total = len(df_new)
        cats = int(df_new.iloc[:, -1].nunique()) if len(df_new.columns) >= 2 else 0
        return jsonify({"success": True, "filename": f.filename, "total_samples": total,
                        "total": total, "categories": cats,
                        "message": f"Dataset de treinamento atualizado com {total} amostras em {cats} categorias. Execute o retreinamento para aplicar."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Upload de Dataset de Avaliação ─────────────────────────────────────────────
@app.route("/upload-eval-dataset", methods=["POST"])
def upload_eval_dataset():
    """Substitui o dataset de AVALIAÇÃO por um novo arquivo .xlsx enviado pelo admin."""
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Apenas arquivos .xlsx são aceitos"}), 400
    try:
        f.save(EVAL_DATASET_PATH)
        df_new = pd.read_excel(EVAL_DATASET_PATH)
        total = len(df_new)
        return jsonify({"success": True, "filename": f.filename, "total_samples": total,
                        "total": total,
                        "message": f"Dataset de avaliação atualizado com {total} amostras."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Treinamento em Tempo Real (SSE) ─────────────────────────────────────────
_train_lock = threading.Lock()
_train_running = False


def _normalize_text(text: str) -> str:
    """Normaliza texto removendo acentos e caracteres especiais."""
    text = str(text).lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


@app.route("/train-stream", methods=["GET"])
def train_stream():
    """
    Endpoint SSE que executa o retreinamento completo do modelo em tempo real,
    emitindo eventos de progresso via Server-Sent Events.
    """
    global pipeline, metrics, _train_running

    if _train_running:
        def already_running():
            yield 'data: {"type":"error","message":"Treinamento já em andamento. Aguarde."}\n\n'
        return Response(stream_with_context(already_running()), mimetype='text/event-stream')

    def generate():
        global pipeline, metrics, _train_running
        _train_running = True
        try:
            import pandas as pd
            from sklearn.pipeline import Pipeline
            from sklearn.feature_extraction.text import TfidfVectorizer
            from sklearn.naive_bayes import MultinomialNB
            from sklearn.model_selection import cross_val_score, StratifiedKFold
            from sklearn.metrics import classification_report, confusion_matrix

            def emit(event_type, **kwargs):
                data = json.dumps({"type": event_type, "ts": datetime.now(timezone.utc).isoformat(), **kwargs})
                return f"data: {data}\n\n"

            yield emit("start", message="Iniciando treinamento do modelo TF-IDF + Naive Bayes...")
            time.sleep(0.1)

            # 1. Carregar dataset
            yield emit("log", step=1, total=8, message=f"Carregando dataset: {TRAIN_DATASET_PATH}")
            time.sleep(0.2)
            df = pd.read_excel(TRAIN_DATASET_PATH)
            n_samples = len(df)
            # Suporte a colunas em português (Categoria/Titulo/Descricao) e inglês (category/title/description)
            col_cat  = 'Categoria' if 'Categoria' in df.columns else 'category'
            col_title = 'Titulo' if 'Titulo' in df.columns else 'title'
            col_desc  = 'Descricao' if 'Descricao' in df.columns else 'description'
            cats = df[col_cat].value_counts().to_dict()
            yield emit("log", step=1, total=8,
                       message=f"Dataset carregado: {n_samples} amostras em {len(cats)} categorias (colunas: {col_cat}, {col_title}, {col_desc})",
                       details=cats)
            time.sleep(0.2)

            # 2. Pré-processamento
            yield emit("log", step=2, total=8, message="Pré-processando textos (normalização, remoção de acentos)...")
            time.sleep(0.3)
            X = (df[col_title].apply(_normalize_text) + ' ' + df[col_desc].apply(_normalize_text)).values
            y = df[col_cat].str.lower().str.replace(' ', '_').values
            unique_cats = sorted(set(y))
            yield emit("log", step=2, total=8,
                       message=f"Pré-processamento concluído: {len(X)} amostras, categorias: {unique_cats}")
            time.sleep(0.2)

            # 3. Construir pipeline TF-IDF
            yield emit("log", step=3, total=8, message="Construindo pipeline TF-IDF (ngram 1-2, max_features=10000)...")
            time.sleep(0.3)
            model = Pipeline([
                ('tfidf', TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=10000)),
                ('clf', MultinomialNB(alpha=0.1))
            ])
            yield emit("log", step=3, total=8,
                       message="Pipeline criado: TfidfVectorizer(ngram=(1,2), max_features=10000) + MultinomialNB(alpha=0.1)")
            time.sleep(0.2)

            # 4. Treinar modelo
            yield emit("log", step=4, total=8, message="Treinando modelo com todos os dados...")
            time.sleep(0.3)
            model.fit(X, y)
            train_acc = model.score(X, y)
            yield emit("progress", step=4, total=8,
                       message=f"Treinamento concluído! Acurácia no treino: {train_acc:.2%}",
                       train_accuracy=round(train_acc, 4))
            time.sleep(0.2)

            # 5. Validação cruzada (5-fold)
            yield emit("log", step=5, total=8, message="Executando validação cruzada 5-fold...")
            time.sleep(0.2)
            cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
            cv_scores = cross_val_score(model, X, y, cv=cv, scoring='accuracy')
            for fold_i, score in enumerate(cv_scores, 1):
                yield emit("fold", step=5, total=8,
                           message=f"Fold {fold_i}/5: {score:.2%}",
                           fold=fold_i, score=round(float(score), 4))
                time.sleep(0.15)
            cv_mean = float(cv_scores.mean())
            cv_std = float(cv_scores.std())
            yield emit("progress", step=5, total=8,
                       message=f"Validação cruzada: {cv_mean:.2%} ± {cv_std:.2%}",
                       cv_mean=round(cv_mean, 4), cv_std=round(cv_std, 4))
            time.sleep(0.2)

            # 6. Avaliação com dataset independente
            yield emit("log", step=6, total=8, message=f"Avaliando com dataset independente: {EVAL_DATASET_PATH}")
            time.sleep(0.3)
            eval_acc = None
            per_category = {}
            try:
                eval_texts, eval_labels = _load_dataset(EVAL_DATASET_PATH)
                eval_preds = model.predict(eval_texts)
                eval_acc = float(np.mean(eval_preds == eval_labels))
                report = classification_report(eval_labels, eval_preds, output_dict=True)
                per_category = {k: {"precision": round(v["precision"], 4),
                                    "recall": round(v["recall"], 4),
                                    "f1_score": round(v["f1-score"], 4),
                                    "support": v["support"]}
                                for k, v in report.items()
                                if k not in ["accuracy", "macro avg", "weighted avg"]}
                cm = confusion_matrix(eval_labels, eval_preds, labels=unique_cats).tolist()
                yield emit("progress", step=6, total=8,
                           message=f"Avaliação independente: {eval_acc:.2%}",
                           eval_accuracy=round(eval_acc, 4),
                           per_category=per_category,
                           confusion_matrix={"labels": unique_cats, "matrix": cm})
            except Exception as e:
                yield emit("warning", step=6, total=8,
                           message=f"Avaliação independente falhou: {e}. Continuando...")
            time.sleep(0.2)

            # 7. Salvar modelo
            yield emit("log", step=7, total=8, message="Salvando modelo treinado em disco...")
            time.sleep(0.3)
            joblib.dump(model, MODEL_PATH)
            yield emit("log", step=7, total=8, message=f"Modelo salvo: {MODEL_PATH}")
            time.sleep(0.2)

            # 8. Atualizar modelo em memória e métricas
            yield emit("log", step=8, total=8, message="Atualizando modelo em memória e métricas...")
            pipeline = model
            now = datetime.now(timezone.utc).isoformat()
            new_metrics = {
                "method": "TF-IDF + Naive Bayes (MultinomialNB)",
                "dataset_size": n_samples,
                "train_accuracy": round(train_acc, 4),
                "cv_accuracy_mean": round(cv_mean, 4),
                "cv_accuracy_std": round(cv_std, 4),
                "categories": unique_cats,
                "category_distribution": {k: int(v) for k, v in cats.items()},
                "last_updated": now,
                "training": {
                    "dataset": os.path.basename(TRAIN_DATASET_PATH),
                    "dataset_size": n_samples,
                    "train_accuracy": round(train_acc, 4),
                    "cv_accuracy_mean": round(cv_mean, 4),
                    "cv_accuracy_std": round(cv_std, 4),
                    "categories": unique_cats,
                    "category_distribution": {k: int(v) for k, v in cats.items()},
                    "new_samples_added": 0,
                    "new_categories": []
                }
            }
            if eval_acc is not None:
                new_metrics["evaluation"] = {
                    "dataset": os.path.basename(EVAL_DATASET_PATH),
                    "dataset_size": len(eval_texts),
                    "eval_accuracy": round(eval_acc, 4),
                    "per_category": per_category,
                    "evaluated_at": now
                }
                new_metrics["eval_accuracy"] = round(eval_acc, 4)
                new_metrics["last_evaluated"] = now
            metrics = new_metrics
            with open(METRICS_PATH, "w", encoding="utf-8") as mf:
                json.dump(metrics, mf, ensure_ascii=False, indent=2)
            yield emit("complete",
                       message="Treinamento concluído com sucesso!",
                       train_accuracy=round(train_acc, 4),
                       cv_accuracy=round(cv_mean, 4),
                       eval_accuracy=round(eval_acc, 4) if eval_acc is not None else None,
                       dataset_size=n_samples,
                       categories=unique_cats)
        except Exception as e:
            yield f'data: {{"type":"error","message":{json.dumps(str(e))}}}\n\n'
        finally:
            _train_running = False

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Access-Control-Allow-Origin': '*'
        }
    )


@app.route("/train-status", methods=["GET"])
def train_status():
    """Verifica se um treinamento está em andamento."""
    return jsonify({"running": _train_running})


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=int(os.environ.get("ML_PORT", 5001)))
    args = parser.parse_args()
    port = args.port
    print(f"[Classifier] Iniciando servidor na porta {port}...")
    print(f"[Classifier] Dataset de TREINAMENTO: {TRAIN_DATASET_PATH}")
    print(f"[Classifier] Dataset de AVALIAÇÃO:   {EVAL_DATASET_PATH}")
    app.run(host="0.0.0.0", port=port, debug=False)
